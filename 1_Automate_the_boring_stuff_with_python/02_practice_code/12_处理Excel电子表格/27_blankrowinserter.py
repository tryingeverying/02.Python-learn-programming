#! python3
# 在指定的文件路径的文件中的N行开始插入M个空行

import openpyxl,sys
from openpyxl.utils import get_column_letter

# line_row = int(sys.argv[1])
# num_row = int(sys.argv[2])
# filename = sys.argv[3]
line_row = int(input('输入要插入的行号'))
num_row = int(input('输入要插入的行数'))
filename = input(r'输入待修改的文件路径')
# 打开文件，获取最大行列序号
ws = openpyxl.load_workbook(filename)
sheet = ws.active

max_row_num , max_column_num = sheet.max_row , sheet.max_column

# 新建工作簿
wb_aim = openpyxl.Workbook()
sheet_aim = wb_aim.active
numfirst = get_column_letter(1) + str(line_row)
num_last = get_column_letter(max_column_num) + str(line_row + num_row - 1)

for i in range(1,max_row_num+1):
    for j in range(1,max_column_num+1):
        if i < line_row:
            sheet_aim.cell(row=i,column=j).value =sheet.cell(row=i,column=j).value
        else:
            sheet_aim.cell(row=i+num_row,column=j).value =sheet.cell(row=i,column=j).value

wb_aim.save(r'F:\Programming\Python\1_Automate_the_boring_stuff_with_python\practice_code\excel file\change_26_multiplicationtable.xlsx')        















