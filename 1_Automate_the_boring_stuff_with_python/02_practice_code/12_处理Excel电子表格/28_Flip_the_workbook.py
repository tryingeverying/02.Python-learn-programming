#! python3
# Flip the workbook.py 实现对工作簿的行转列，列转行

import openpyxl
from openpyxl.utils import get_column_letter

file_path = input(r'请输入要翻转的文件路径')

# 获取待处理工作簿的最大行号和列号
wb = openpyxl.load_workbook(file_path)
sheet = wb.active
max_row_num ,max_column_num = sheet.max_row,sheet.max_column

# 将翻转后的内容写入新工作簿
wb_new = openpyxl.Workbook()
sheet_new = wb_new.active

for i in range(1,max_row_num+1):
    for j in range(1,max_column_num+1):
        sheet_new.cell(row = j,column=i).value = sheet.cell(row = i,column=j).value

wb_new.save(r'F:\Programming\Python\1_Automate_the_boring_stuff_with_python\practice_code\excel file\new_example.xlsx')













