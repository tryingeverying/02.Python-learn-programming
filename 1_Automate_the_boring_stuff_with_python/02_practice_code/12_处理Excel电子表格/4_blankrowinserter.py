#! python3
# 4_blankrowinserter.py 在指定文件的指定行插入指定行

import openpyxl
from openpyxl.utils import get_column_letter

file_path = input(r'请输入文件路径')
insert_line = int(input('从第几行插入？'))
line_num = int(input('插入多少行？'))

wb = openpyxl.load_workbook(file_path)
sheet = wb.active

max_row_num , max_column_num = sheet.max_row , sheet.max_column
wb_aim = openpyxl.Workbook()
sheet_aim = wb_aim.active
first_num = get_column_letter(1)+str(insert_line)
last_num = get_column_letter(max_column_num)+str(insert_line+line_num-1)

for i in range(1,max_row_num+1):
    for j in range(1,max_column_num+1):
        if i < insert_line:
            sheet_aim.cell(row=i,column=j).value = sheet.cell(row=i,column=j).value 
        else:
            sheet_aim.cell(row=i+line_num,column=j).value = sheet.cell(row=i,column=j).value 

# sheet_aim.merge_cells (''+first_num+':'+last_num+'')
wb_aim.save(r'F:\Programming\Python\1_Automate_the_boring_stuff_with_python\Reproduce_code\12_处理Excel电子表格\4_result.xlsx')















