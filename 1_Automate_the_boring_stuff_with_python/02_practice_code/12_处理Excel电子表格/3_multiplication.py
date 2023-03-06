#! python3
# 3_multiplication.py 生成指定的乘法表

import openpyxl
from openpyxl.styles import Font

num = int(input('请输入一个数字'))

wb = openpyxl.Workbook()
sheet = wb.active

boldfont = Font(bold=True)

for i in range(1,num+1):
    sheet.cell(row=1,column=i+1).value = i
    sheet.cell(row=1,column=i+1).font = boldfont
for j in range(1,num+1):
    sheet.cell(row=j+1,column=1).value = j
    sheet.cell(row=j+1,column=1).font = boldfont


for i in range(2,num+2):
    for j in range(2,num+2):
        sheet.cell(row=i,column=j).value = sheet.cell(row=i,column=1).value *sheet.cell(row=1,column=j).value

wb.save(r'F:\Programming\Python\1_Automate_the_boring_stuff_with_python\Reproduce_code\12_处理Excel电子表格\3_result.xlsx')
print('done')



















