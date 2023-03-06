#! python3
# 实现Excel的行列转置

import openpyxl

file_path = input(r'请输入待处理的文件路径')
#读取文件，并获取最大行号列号
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

max_row_num ,max_column_num = sheet.max_row , sheet.max_column

#新建一个储存转置后内容的文件
wb_new = openpyxl.Workbook()
sheet_new = wb_new.active

# 实现内容转置
for i in range(1,max_row_num+1):
    for j in range(1,max_row_num+1):
        sheet_new.cell(row=j,column=i).value = sheet.cell(row=i,column=j).value

wb_new.save(r'F:\Programming\Python\1_Automate_the_boring_stuff_with_python\Reproduce_code\12_处理Excel电子表格\5_result.xlsx')






























